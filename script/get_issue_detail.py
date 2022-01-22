import requests
from bs4 import BeautifulSoup
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import load_workbook
import re
import csv

def make_soup(issue_url_input):
    article_res = requests.get(issue_url_input)
    return BeautifulSoup(article_res.content, 'html.parser')

def get_issue_title(soap_article):
    return soap_article.find('h1', class_ = 'title').get_text()

def get_abstract(soap_article):
    if soap_article.find('section', class_ = 'abstract'):
        return soap_article.find('section', class_ = 'abstract').get_text().replace('\t', "").replace('Abstract\n', "")
    else:
        return 'Abstract not found'

def not_ascii(raw_abstract):
    not_ascii_chars = ''
    for char in raw_abstract:
        if not char.isascii():
            not_ascii_chars += f' {char}'
    return not_ascii_chars

def clean_abstract(raw_abstract):
    if raw_abstract:
        cleaned_abstract = raw_abstract.replace("\r\n", " ").replace("\n","")
        cleaned_abstract = ILLEGAL_CHARACTERS_RE.sub(r'', cleaned_abstract)
        return cleaned_abstract
    else:
        return 'Abstract not available'

def get_authors(soap_article):
    author_list = []
    article_authors = soap_article.find_all('li', class_ ='author')
    for author in article_authors:
        author_text = author.text.strip()
        author_list.append(author_text)
    return author_list

def get_volume_issue(soap_article):
    return soap_article.find_all('li', class_ = 'journal')[1].text.strip().replace('\n', "").replace('\t', "")
    

def get_issue_urls_from_csv(filename):
    issue_url_list=[]         #an empty list to store the urls
    with open(filename, 'r') as journal_csv:
        reader = csv.reader(journal_csv, delimiter=',')
        for row in reader:
            issue_url_list.append(row[0])
    return issue_url_list

def write_excel(filename,start_row, value_to_write):
    wb = load_workbook(filename)
    ws = wb.active
    for i in range(0,len(value_to_write)):
        ws.cell(row = start_row, column = i+1, value = value_to_write[i] )
    wb.save(filename)
    
#getting issue urls from csv file populated before
#load into list
urls = get_issue_urls_from_csv('issue_record.csv')

excel_row = 2
for i in range(0, len(urls)):
    print(f"Processing record {i+1} of {len(urls)} ({ round ((100 * (i+1) / len(urls)),2)}%)... ")

    soap_issue = make_soup(urls[i])
    print(f"Scrapping {urls[i]}" , sep ='')
    issue_title = get_issue_title(soap_issue)
    print(f"Journal Title {issue_title}")

    raw_abstract  = get_abstract(soap_issue)
    non_ascii = not_ascii(raw_abstract)
    cleaened_abstract = clean_abstract(raw_abstract)

    vlm_issue = get_volume_issue(soap_issue)
    authors = get_authors(soap_issue)

    row_values = [urls[i], issue_title, vlm_issue, cleaened_abstract, non_ascii]
    for author in authors:
        row_values.append(author)

    print(f'writing row {excel_row} ..')
    write_excel('journal_all_reformat.xlsx', excel_row, row_values)
    excel_row += 1
    



